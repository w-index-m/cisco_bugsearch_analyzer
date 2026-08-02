"""
Cisco Bug Search Analyzer - コアロジック

Streamlit (app.py) と CLI (cli.py) の両方から利用する共通処理。
Streamlit に依存しないため、エージェントやスクリプトから直接 import して使える。
"""
import io
import re
import json
import time
from html import unescape
from datetime import datetime, date as _date
from functools import wraps

import pandas as pd
import requests
import yaml
from deep_translator import GoogleTranslator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import deepl
    DEEPL_AVAILABLE = True
except Exception:
    DEEPL_AVAILABLE = False

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except Exception:
    GEMINI_AVAILABLE = False

try:
    from groq import Groq
    GROQ_AVAILABLE = True
except Exception:
    GROQ_AVAILABLE = False


# ==================== データ読み込み ====================

def load_csv(file_path):
    """CSV ファイルを読み込み、カラム名の前後空白を除去"""
    df = pd.read_csv(file_path)
    df.columns = df.columns.str.strip()
    return df


# Cisco Bug Search のエクスポートに必要な標準列名。この名前で他の処理が参照するため、
# アップロードされたファイルの列名がこれと違う場合は normalize_bug_columns() で正規化する
REQUIRED_BUG_COLUMNS = [
    "BUG Id", "BUG headline", "URL", "Bug Status", "Bug Severity",
    "Known Fixed Releases", "Last Modified", "Product - Series",
    "Known Affected Release(s)", "Release Note Enclosure",
]

# 製品ライン（Catalyst/Nexus等）によってエクスポートの列名が微妙に異なることがある
# （例: Nexus では "Bug Status"/"Bug Severity" ではなく単に "Status"/"Severity"）ため、
# よくある別名から標準列名への変換テーブル
_COLUMN_ALIASES = {
    "bug id": "BUG Id",
    "bug headline": "BUG headline",
    "headline": "BUG headline",
    "url": "URL",
    "bug status": "Bug Status",
    "status": "Bug Status",
    "bug severity": "Bug Severity",
    "severity": "Bug Severity",
    "known fixed releases": "Known Fixed Releases",
    "known fixed": "Known Fixed Releases",
    "known affected release(s)": "Known Affected Release(s)",
    "known affected releases": "Known Affected Release(s)",
    "known affected": "Known Affected Release(s)",
    "last modified": "Last Modified",
    "product - series": "Product - Series",
    "product": "Product - Series",
    "release note enclosure": "Release Note Enclosure",
}


def normalize_bug_columns(df):
    """
    アップロードされたバグ一覧の列名を Cisco Bug Search の標準形式に正規化する。

    製品ライン（Nexus等）によっては "Bug Status"/"Bug Severity" ではなく
    "Status"/"Severity" のように短縮された列名でエクスポートされることがあるため、
    既知の別名（_COLUMN_ALIASES）を介して標準列名へリネームする。
    標準列がリネーム後も見つからない場合は、空文字列の列として補い、
    以降の処理（列への直接アクセス）で KeyError にならないようにする。
    """
    df = df.copy()
    lower_cols = {str(c).strip().lower(): c for c in df.columns}
    rename_map = {}
    for alias, canonical in _COLUMN_ALIASES.items():
        if canonical in df.columns:
            continue
        if alias in lower_cols:
            rename_map[lower_cols[alias]] = canonical
    df = df.rename(columns=rename_map)

    for col in REQUIRED_BUG_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    return df


# ==================== リリースノート処理 ====================

def clean_html_tags(text):
    """HTML タグを削除して日本語対応テキストに変換（NaN/非文字列も安全に空文字として扱う）"""
    if not isinstance(text, str) or not text:
        return ""
    text = re.sub(r'<[^>]+>', '', text)
    text = unescape(text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


_RELEASE_NOTE_SECTION_MAPPING = {
    'Symptom': '症状',
    'Symptôme': '症状',
    'Conditions': '条件',
    "Conditions d'activation": '条件',
    'Workaround': '回避策',
    'Contournement': '回避策',
    'Further Problem Description': '詳細説明',
    'Description additionnelle du problème': '詳細説明',
}

# 長いキーワードから先にマッチさせる（"Conditions d'activation" が "Conditions" に
# 食われて誤分割されるのを防ぐため）
_RELEASE_NOTE_KEYWORDS_SORTED = sorted(_RELEASE_NOTE_SECTION_MAPPING.keys(), key=len, reverse=True)
_RELEASE_NOTE_SPLIT_RE = re.compile(
    r'(' + '|'.join(re.escape(k) for k in _RELEASE_NOTE_KEYWORDS_SORTED) + r')[:\s]*',
    re.IGNORECASE,
)


def parse_release_note(note_text):
    """
    リリースノートをセクション別（症状/条件/回避策/詳細説明）に解析する。

    clean_html_tags() が改行を含む空白を全てスペースに正規化するため、
    各セクションの本文は「次のセクション見出しが現れる直前まで」を境界として
    切り出す（固定文字数での打ち切りは、次の見出し文字列がそのまま本文に
    混入してしまうため使わない）。
    """
    text = clean_html_tags(note_text)
    if not text:
        return {}

    sections = {}

    matches = list(_RELEASE_NOTE_SPLIT_RE.finditer(text))

    for i, m in enumerate(matches):
        jp_key = next(
            (v for k, v in _RELEASE_NOTE_SECTION_MAPPING.items() if k.lower() == m.group(1).lower()),
            None,
        )
        if jp_key is None:
            continue

        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        content = text[start:end].strip()

        if content and jp_key not in sections:
            sections[jp_key] = content

    return sections


def get_cisco_release_notes_url(product, version):
    """Cisco 公式リリースノート URL を生成"""
    product_lower = product.lower()
    if 'catalyst 9200' in product_lower or '9200' in product_lower:
        version_short = version.replace('.', '-')
        return f"https://www.cisco.com/c/en/us/td/docs/switches/lan/catalyst9200/software/release/{version_short}/release_notes/ol-{version_short}-9200.html"
    elif 'catalyst 9300' in product_lower or '9300' in product_lower:
        version_short = version.replace('.', '-')
        return f"https://www.cisco.com/c/en/us/td/docs/switches/lan/catalyst9300/software/release/{version_short}/release_notes/ol-{version_short}-9300.html"
    elif 'catalyst 9400' in product_lower or '9400' in product_lower:
        version_short = version.replace('.', '-')
        return f"https://www.cisco.com/c/en/us/td/docs/switches/lan/catalyst9400/software/release/{version_short}/release_notes/ol-{version_short}-9400.html"
    else:
        return None


# ==================== 翻訳 ====================

def _cache_success_only(fn):
    """
    lru_cache と違い、翻訳の失敗（None）はキャッシュしない。

    無料の翻訳エンドポイントは一時的なレート制限等で稀に失敗するが、
    lru_cache でそのまま失敗結果をキャッシュすると、同じ見出しが二度と
    再翻訳されず永久に英語のまま表示され続けてしまう
    （実際にこの不具合が発生し、複数バグのヘッドラインが一括で
    未翻訳になる原因になっていた）。成功した結果のみキャッシュすることで、
    次回呼び出し時に再試行のチャンスを残す。
    """
    cache = {}

    @wraps(fn)
    def wrapper(*args):
        if args in cache:
            return cache[args]
        result = fn(*args)
        if result:
            cache[args] = result
        return result

    return wrapper


@_cache_success_only
def translate_headline_deepl(text, api_key):
    """DeepL API を使用して日本語に翻訳"""
    if not text or len(text) < 3:
        return text
    try:
        translator = deepl.Translator(api_key)
        result = translator.translate_text(text, source_lang="EN", target_lang="JA")
        return result.text
    except Exception:
        return None


@_cache_success_only
def translate_headline_google(text):
    """
    Google Translate を使用して日本語に翻訳する。

    無料エンドポイントは一時的なレート制限等で稀に失敗することがあるため、
    短い間隔を空けて最大3回まで再試行する。
    """
    if not text or len(text) < 3:
        return text
    for attempt in range(3):
        try:
            translator = GoogleTranslator(source='en', target='ja')
            return translator.translate(text)
        except Exception:
            if attempt < 2:
                time.sleep(0.5 * (attempt + 1))
    return None


def translate_headline_nvidia(text, api_key):
    """NVIDIA Riva Translation を使用して日本語に翻訳"""
    if not text or len(text) < 3:
        return None
    if not api_key:
        return None

    try:
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": "nvidia/riva-translate-4b-instruct-v2",
            "messages": [
                {
                    "role": "user",
                    "content": f"Translate the following text from English to Japanese. Only return the translated text without any explanation.\n\nText: {text}"
                }
            ],
            "temperature": 0.5,
            "top_p": 0.9,
            "max_tokens": 512
        }

        response = requests.post(
            "https://integrate.api.nvidia.com/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=10
        )

        if response.status_code == 200:
            result = response.json()
            if "choices" in result and len(result["choices"]) > 0:
                return result["choices"][0]["message"]["content"].strip()
        return None

    except Exception:
        return None


def translate_headline(text, engine='google', deepl_api_key=None, nvidia_api_key=None):
    """翻訳エンジンを指定してヘッドラインを翻訳（フォールバック: 指定エンジン → Google）"""
    if not text or len(text) < 3:
        return text

    if engine == 'nvidia' and nvidia_api_key:
        result = translate_headline_nvidia(text, nvidia_api_key)
        if result:
            return result

    if engine == 'deepl' and deepl_api_key and DEEPL_AVAILABLE:
        result = translate_headline_deepl(text, deepl_api_key)
        if result:
            return result

    result = translate_headline_google(text)
    if result:
        return result

    return text


# ==================== 分析データの永続化 ====================

def save_analysis_to_json(analysis_data):
    """分析結果を JSON 文字列に変換"""
    output = {
        "timestamp": datetime.now().isoformat(),
        "analysis_count": len(analysis_data),
        "bugs": []
    }

    for bug_id, data in analysis_data.items():
        output["bugs"].append({
            "bug_id": bug_id,
            "possibility": data.get("possibility", "-"),
            "tags": data.get("tags", []),
            "comment": data.get("comment", "")
        })

    return json.dumps(output, ensure_ascii=False, indent=2)


def load_analysis_from_json(json_str):
    """JSON 文字列から分析結果を読み込み。失敗時は空 dict を返す"""
    try:
        data = json.loads(json_str)
        analysis = {}
        for bug in data.get("bugs", []):
            analysis[bug["bug_id"]] = {
                "possibility": bug.get("possibility", "Medium"),
                "tags": bug.get("tags", []),
                "comment": bug.get("comment", "")
            }
        return analysis
    except Exception:
        return {}


# ==================== AI による発生可能性判定 ====================

def _build_assessment_prompt(headline, release_note, user_comment):
    return f"""以下のバグ情報から、発生の可能性を High/Medium/Low で判定してください。

【バグタイトル（日本語）】
{headline}

【リリースノート】
{release_note[:500] if release_note else ""}

【ユーザーコメント】
{user_comment if user_comment else "（コメントなし）"}

【判定フォーマット】
可能性: [High/Medium/Low]
理由: [簡潔な理由]

日本語で回答してください。"""


def assess_bug_with_groq(headline, release_note, user_comment, api_key):
    """Groq API を使用して発生可能性を判定"""
    if not api_key or not GROQ_AVAILABLE:
        return None

    try:
        client = Groq(api_key=api_key)
        prompt = _build_assessment_prompt(headline, release_note, user_comment)

        message = client.chat.completions.create(
            model="mixtral-8x7b-32768",
            max_tokens=200,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        return message.choices[0].message.content
    except Exception:
        return None


def assess_bug_with_gemini(headline, release_note, user_comment, api_key):
    """Gemini API を使用して発生可能性を判定"""
    if not api_key or not GEMINI_AVAILABLE:
        return None

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-pro')
        prompt = _build_assessment_prompt(headline, release_note, user_comment)

        response = model.generate_content(prompt)
        return response.text
    except Exception:
        return None


def assess_bug_with_open_router(headline, release_note, user_comment, api_key):
    """Open Router API を使用して発生可能性を判定"""
    if not api_key:
        return None

    try:
        prompt = _build_assessment_prompt(headline, release_note, user_comment)

        response = requests.post(
            url="https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "HTTP-Referer": "https://streamlit.io",
                "X-Title": "Cisco Bug Search Analyzer"
            },
            json={
                "model": "mistralai/mistral-7b-instruct",
                "messages": [
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 200
            },
            timeout=10
        )

        if response.status_code == 200:
            result = response.json()
            if "choices" in result and len(result["choices"]) > 0:
                return result["choices"][0]["message"]["content"]
        return None
    except Exception:
        return None


def assess_bug_possibility(headline_ja, release_note, user_comment, groq_key=None, gemini_key=None, open_router_key=None):
    """
    複数の AI を試してバグの発生可能性を判定

    フォールバック順: Groq → Gemini → Open Router
    """
    assessments = [
        ("Groq", assess_bug_with_groq(headline_ja, release_note, user_comment, groq_key)),
        ("Gemini", assess_bug_with_gemini(headline_ja, release_note, user_comment, gemini_key)),
        ("Open Router", assess_bug_with_open_router(headline_ja, release_note, user_comment, open_router_key))
    ]

    for engine_name, result in assessments:
        if result:
            return {"engine": engine_name, "result": result}

    return None


# ==================== AI による技術文の要約 ====================
# リリースノートの症状/回避策/詳細説明には、16進アドレスやスタックトレース、
# プロセスダンプ等の生ログがそのまま含まれることが多く、翻訳しただけでは
# レポートとして冗長になりがち。AI（Groq → Gemini → Open Router の順で試行、
# いずれも「AI 分析エンジン」に設定済みのキーを流用）で日本語の要点のみに要約する。
# キー未設定の場合は None を返し、呼び出し側は通常の翻訳結果にフォールバックする。

_SUMMARY_MAX_INPUT_CHARS = 2000


def _build_summary_prompt(text):
    return (
        "以下はCisco機器のバグに関する技術的な説明文（英語）です。"
        "16進数のメモリアドレス・スタックトレース・プロセスダンプ等の細かいログ情報は"
        "無視し、実際に何が起きるかという要点だけを日本語で1〜2文（80文字程度）に"
        "要約してください。要約文以外は出力しないでください。\n\n"
        f"{text[:_SUMMARY_MAX_INPUT_CHARS]}"
    )


def summarize_with_groq(text, api_key):
    if not api_key or not GROQ_AVAILABLE:
        return None
    try:
        client = Groq(api_key=api_key)
        message = client.chat.completions.create(
            model="mixtral-8x7b-32768",
            max_tokens=150,
            messages=[{"role": "user", "content": _build_summary_prompt(text)}]
        )
        return message.choices[0].message.content.strip()
    except Exception:
        return None


def summarize_with_gemini(text, api_key):
    if not api_key or not GEMINI_AVAILABLE:
        return None
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-pro')
        response = model.generate_content(_build_summary_prompt(text))
        return response.text.strip()
    except Exception:
        return None


def summarize_with_open_router(text, api_key):
    if not api_key:
        return None
    try:
        response = requests.post(
            url="https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "HTTP-Referer": "https://streamlit.io",
                "X-Title": "Cisco Bug Search Analyzer"
            },
            json={
                "model": "mistralai/mistral-7b-instruct",
                "messages": [{"role": "user", "content": _build_summary_prompt(text)}],
                "max_tokens": 150
            },
            timeout=15
        )
        if response.status_code == 200:
            result = response.json()
            if "choices" in result and len(result["choices"]) > 0:
                return result["choices"][0]["message"]["content"].strip()
        return None
    except Exception:
        return None


def summarize_technical_text_ja(text, groq_key=None, gemini_key=None, open_router_key=None):
    """
    技術的な説明文（英語）を、AI で日本語の要点のみに要約する。

    フォールバック順: Groq → Gemini → Open Router。いずれのキーも無い、または
    いずれも失敗した場合は None を返す（呼び出し側で通常の翻訳結果にフォールバック
    すること）。短い文（100文字未満）は要約の必要が薄いため None を返す。
    """
    if not text or len(text) < 100:
        return None

    for summarizer, key in (
        (summarize_with_groq, groq_key),
        (summarize_with_gemini, gemini_key),
        (summarize_with_open_router, open_router_key),
    ):
        result = summarizer(text, key)
        if result:
            return result

    return None


# ==================== バージョン比較 ====================

_VERSION_DIGITS_RE = re.compile(r'\d+')


def _parse_version_tuple(token):
    """
    バージョン文字列から (major, minor, patch) の整数タプルを抽出する。

    コードネーム接頭辞（Dublin-17.12.5）、括弧表記（17.12(4.2)）、
    末尾の英字サフィックス（17.12.4a）、旧 IOS 表記（3.10.1E_ngwc）などを許容する。
    数値が2つ未満しか取れない場合（sdk-master など）は None を返す。
    """
    digits = _VERSION_DIGITS_RE.findall(str(token))
    if len(digits) < 2:
        return None
    nums = [int(d) for d in digits[:3]]
    while len(nums) < 3:
        nums.append(0)
    return tuple(nums)


def version_affects_bug(affected_str, fixed_str, target_version):
    """
    target_version 時点でバグが「影響あり（未修正）」かどうかを判定する。

    ロジック:
    1. target と同じトレイン（major.minor）の Known Affected Release(s) の中に
       target 以前（同じか古い）のバージョンがあれば、そのトレインで「影響が始まっている」とみなす
    2. 同じトレインの Known Fixed Releases の中に target 以前（同じか古い）の修正版が
       無ければ「まだ直っていない」と判定し True を返す
    3. target_version が数値として解析できない場合（例: "master" 等）は、
       従来通り Known Affected Release(s) への部分文字列一致にフォールバックする

    注意（推測ロジックである点）:
    これは「Known Affected Release(s) に載っている最も古いバージョンから、
    修正版が出るまでの間はずっと影響が続いている」という前提に基づく推測であり、
    CSVデータ上の裏付けではない。例えば 17.15.2 のみが Known Affected Release(s)
    に記載されていて 17.15.5 で本関数が True を返した場合でも、実際には中間の
    17.15.3 や 17.15.4 では別の変更により症状が出なくなっている可能性があり、
    その逆に 17.15.4 では発生しないが 17.15.5 で再発している可能性もある。
    「発生する可能性がある」という注意喚起として使い、確定情報としては
    Cisco Bug Search の該当バグページやリリースノートで必ず裏取りすること。
    """
    target = _parse_version_tuple(target_version)

    if target is None:
        return str(target_version).lower() in str(affected_str).lower()

    train = target[:2]

    affected_tokens = str(affected_str).split() if pd.notna(affected_str) else []
    affected_versions = [
        v for v in (_parse_version_tuple(t) for t in affected_tokens)
        if v is not None and v[:2] == train
    ]

    if not affected_versions:
        return False

    if not any(v <= target for v in affected_versions):
        # 同トレインの影響開始バージョンが target より後 → まだそのバージョンに到達していない
        return False

    fixed_tokens = str(fixed_str).split() if pd.notna(fixed_str) else []
    fixed_versions = [
        v for v in (_parse_version_tuple(t) for t in fixed_tokens)
        if v is not None and v[:2] == train
    ]

    if any(v <= target for v in fixed_versions):
        # 同トレインで target 以前に修正版がリリース済み → 既に直っている
        return False

    return True


# ==================== バグの自動分類・発生しやすさ推定 ====================
# ユーザーが手元で使っていた集計スクリプト（Nexus/Catalyst 向け）のキーワード分類・
# 発生しやすさ推定ロジックを移植し、製品を問わず使えるよう分類ルールを統合したもの。
# いずれもヘッドライン中のキーワードマッチによる目安であり、実際のヒット件数等の
# 統計データに基づくものではない。

_BUG_FEATURE_RULES = [
    ("IPsec/VPN", r"\bipsec\b|\bvpn\b|\bike\b|\bl2tp\b|\bpptp\b|dmvpn"),
    ("vPC/ピアリンク", r"\bvpc\b|peer[- ]?link|peer[- ]?keepalive|peer[- ]?gateway"),
    ("VXLAN/EVPN", r"vxlan|evpn|\bnve\b|vtep|type-5|type-2|anycast gateway|l3vni|l2vni"),
    ("BGP", r"\bbgp\b"),
    ("OSPF", r"\bospf\b|ospfv3"),
    ("EIGRP", r"\beigrp\b"),
    ("IS-IS", r"\bisis\b|is-is"),
    ("PIM/マルチキャスト", r"\bpim\b|multicast|igmp|msdp|\bmroute"),
    ("HSRP/VRRP/冗長化/スタック", r"\bhsrp\b|\bvrrp\b|\bglbp\b|first hop|fhrp|redundan|stackwise|\bsvl\b|\bstack\b|switchover"),
    ("STP/L2/MACアドレス学習", r"spanning[- ]?tree|\bstp\b|\bmst\b|bpdu|\bvlan\b|mac address|\bl2\b"),
    ("LACP/EtherChannel", r"port[- ]?channel|\blacp\b|\blag\b|ether[- ]?channel"),
    ("FEX", r"\bfex\b|fabric extender|satellite"),
    ("OTV", r"\botv\b|overlay transport"),
    ("MPLS/SR", r"\bmpls\b|segment routing|\bsr-?te\b|\bldp\b|\bl3vpn\b"),
    ("QoS", r"\bqos\b|policy[- ]?map|class[- ]?map|policer|\bdscp\b"),
    ("TrustSec(CTS/SGACL)", r"\bcts\b|sgacl|trustsec|security[- ]?group"),
    ("FCoE/ストレージ", r"\bfcoe\b|fibre channel|\bfc\b san|\bnpv\b|\bnpiv\b"),
    ("SNMP", r"\bsnmp\b"),
    ("NTP/PTP", r"\bntp\b|\bptp\b|precision time"),
    ("Syslog/ログ", r"syslog|logging\b|\blog\b"),
    ("SD-WAN/クラウド", r"sd-?wan|vmanage|vsmart|vbond|\bgcp\b|\bazure\b|\bnsi\b|gvnic|onboarding"),
    ("NX-API/プログラマビリティ", r"nx-?api|netconf|restconf|\byang\b|grpc|gnmi|telemetry|\bapi\b"),
    ("DHCP/IPアドレス", r"\bdhcp\b|dhcpv6|ip address|\barp\b|\bnd\b"),
    ("SPAN/ERSPAN/監視", r"\bspan\b|erspan|sflow|netflow|\bmirror\b"),
    ("認証(AAA/802.1X)", r"\bmacsec\b|\bmka\b|\bdot1x\b|802\.1x|\baaa\b|tacacs|radius|radsec|control plane|\bcopp\b"),
    ("無線(WLC/AP)", r"\bwlc\b|wireless|access point|\bap\b joined|\bcapwap\b"),
    ("ライセンス/システム/リソース", r"licens|\bcpu\b|memory|\bmem\b|\bprocess\b|kernel|\bcore\b|resource|\bdisk\b"),
    ("ハードウェア/転送(ASIC/FED/OEF)", r"\basic\b|forwarding|\bfib\b|\btcam\b|\bfed\b|\boef\b|hardware|linecard|line card|\bsup\b|module|transceiver|\bsfp\b|\bqsfp\b|\bpoe\b"),
    ("ISSU/インストール/ブート", r"\bissu\b|install|upgrade|downgrade|\bboot\b|bios|\bepld\b|\bimage\b"),
    ("CLI/管理", r"\bcli\b|show tech|show command|management|webui|web gui"),
]

_BUG_SYMPTOM_RULES = [
    ("脆弱性(DoS/セキュリティ)", r"vulnerabilit|\bcve\b|denial of service|\bdos\b|exploit|security advisory|unauthenticated"),
    ("クラッシュ", r"crash|\bcore\b|coredump|core file|segfault|signal 11|sigabrt"),
    ("予期しないリロード/再起動", r"unexpected(ly)? (reload|reboot|restart)|reset(s)? unexpected|reload(s)? unexpected"),
    ("リロード/再起動", r"\breload|\breboot|restart|reset\b"),
    ("メモリリーク", r"memory leak|mem leak|\bleak\b"),
    ("高CPU", r"high cpu|cpu (spike|util|high)|100% cpu"),
    ("ハング/無応答", r"hang|unresponsive|stuck|freeze|deadlock|not respond"),
    ("フラップ/不安定", r"flap|unstable|bounce"),
    ("トレースバック", r"traceback|back[- ]?trace"),
    ("パケットドロップ/損失", r"packet (drop|loss)|drop(s|ped|ping)?\b|black[- ]?hole|traffic loss"),
    ("誤動作/表示", r"incorrect|wrong|mismatch|inconsistent|stale|not update|does not (show|display|reflect)"),
    ("誤動作/失敗", r"fail|error|unable|not work|does not work|not function|broken"),
]


def classify_bug_feature(text):
    """ヘッドラインのキーワードから利用機能カテゴリを推定する（最大2件、"/"区切り）"""
    if not text:
        return "その他/システム"
    hits = [name for name, pattern in _BUG_FEATURE_RULES if re.search(pattern, text, re.IGNORECASE)]
    return " / ".join(hits[:2]) if hits else "その他/システム"


def classify_bug_symptom(text):
    """ヘッドラインのキーワードから症状カテゴリ（素因）を推定する（最大2件、"/"区切り）"""
    if not text:
        return "その他"
    hits = [name for name, pattern in _BUG_SYMPTOM_RULES if re.search(pattern, text, re.IGNORECASE)]
    return " / ".join(hits[:2]) if hits else "その他"


_OCCURRENCE_RARE_KEYWORDS = [
    "multiple", "repeated", "aggressive", "stress", "scale", "scaled", "race",
    "corner", "rare", "intermittent", "specific", "continuous", "flap",
    "churn", "soak", "overnight", "thousand", "back-to-back", "iteration",
    "toggl", "burst", "bulk", "endlessly", "high scale", "under load",
]
_OCCURRENCE_COMMON_KEYWORDS = [
    "memory leak", "leak", "on reload", "after upgrade", "after reboot",
    "on boot", "at boot", "on startup", "every ", "always", "over time",
    "gradually", "high cpu",
]


def estimate_occurrence_likelihood(status, headline, target_affected=None):
    """
    バグの発生しやすさを Bug Status とヘッドラインのキーワードから推定する。

    実際のヒット件数や統計データに基づくものではなく、あくまで目安。重要な判断の
    前には必ずバグページ本文で実際の発生条件を確認すること。

    Args:
        status: Bug Status（Open/Fixed/Duplicate/Terminated 等）
        headline: BUG headline（英語原文）
        target_affected: 指定バージョンへの影響有無（True/False）。None なら注記なし
    """
    s = (status or "").strip().lower()
    h = (headline or "").lower()
    tail = "（※指定バージョンでは影響なし・参考情報）" if target_affected is False else ""

    if s == "duplicate":
        return "懸念低: 別バグに統合(Duplicate)＝実質クローズ" + tail
    if s == "terminated":
        return "懸念低: 打ち切り(Terminated)" + tail
    if s == "unreproducible":
        return "懸念低: 再現不可(Unreproducible)＝発生性低い" + tail
    if s in ("fixed", "resolved", "closed", "verified"):
        return "修正済: 該当リリースで解消(懸念なし)" + tail

    if re.search(r"vulnerabilit|denial of service|\bdos\b|unauthenticated", h):
        return "中〜高（未修正/脆弱性）: 設定・攻撃条件が揃えば発生し得る。SMU/修正版の早期適用を推奨" + tail

    rare = next((k for k in _OCCURRENCE_RARE_KEYWORDS if k in h), None)
    common = next((k for k in _OCCURRENCE_COMMON_KEYWORDS if k in h), None)
    if common and not rare:
        return f"中〜高（未修正）: 通常操作/時間経過で発生し得る（条件: {common.strip()}）" + tail
    if rare:
        return f"低〜中（未修正）: 特定条件/反復・高負荷時に発生（条件: {rare}）＝通常運用では起きにくい" + tail
    return "中（未修正）: 明確な高頻度要因は少。発生条件は原文/URL参照" + tail


# ==================== バグ検索 ====================

_FEATURE_QUOTED_RE = re.compile(r'["“”](.+?)["“”]')
_FEATURE_DELIMITER_RE = re.compile(r'[,，、\s]+')


def split_feature_terms(feature):
    """
    機能検索文字列を複数のキーワードに分割する。

    区切り文字: 全角/半角カンマ（, ， 、）およびスペース（全角/半角）。
    「"Catalyst 9300"」のようにダブルクォート（全角/半角）で囲むと、
    中にスペースが含まれていても1つのキーワードとして保持される。

    例:
        "VPN Multicast"       -> ["VPN", "Multicast"]（スペース区切りでOR）
        "VPN,Multicast"       -> ["VPN", "Multicast"]（カンマ区切りでOR）
        '"Catalyst 9300" VPN' -> ["Catalyst 9300", "VPN"]（クォート部分は1語のまま）
        "Catalyst 9300"（クォート無し） -> ["Catalyst", "9300"]（スペースで分割される）
    """
    if not feature:
        return []

    terms = [m.group(1).strip() for m in _FEATURE_QUOTED_RE.finditer(feature) if m.group(1).strip()]
    remainder = _FEATURE_QUOTED_RE.sub(' ', feature)
    terms += [t.strip() for t in _FEATURE_DELIMITER_RE.split(remainder) if t.strip()]

    return terms


def search_bugs(df, feature=None, version=None, severity=None, ios_version=None, sort_by=None):
    """
    条件に応じて df をフィルタリングする

    Args:
        df: バグデータ全体
        feature: Product - Series / BUG headline に対する部分一致。
            カンマ（, ， 、）またはスペース（全角/半角）区切りで複数キーワードを指定すると
            OR 検索になる（例: "VPN Multicast" や "VPN,Multicast" はどちらも同じ）。
            スペースを含む語をそのまま1語として検索したい場合はダブルクォートで囲む
            （例: '"Catalyst 9300" VPN'）。split_feature_terms 参照。
        version: 検索バージョン。同トレイン内で「これ以前のバージョンから影響していて、
            まだ修正版が出ていない」バグも含めてマッチする（version_affects_bug 参照）。
            数値として解釈できない文字列を渡した場合は部分一致にフォールバックする。
            ※ これは推測ロジックであり、中間バージョン（例: 17.15.2 のみ記載の場合の
            17.15.3 や 17.15.4）で実際に発生するかどうかを保証するものではない。
            「発生する可能性がある」候補として扱い、確定判断は個別バグページで確認すること。
        severity: 対象とする Severity のリスト（例: [1, 2, 3]）
        ios_version: 特定 IOS バージョンでの絞り込み（version と併用可、判定ロジックは同じ）
        sort_by: "severity" | "last_modified" | "bug_id"

    Returns:
        絞り込み・ソート済みの DataFrame
    """
    results = df.copy()

    if severity:
        results = results[results["Bug Severity"].astype(int).isin(severity)]

    if feature:
        terms = split_feature_terms(feature)
        if terms:
            mask = pd.Series(False, index=results.index)
            for term in terms:
                mask = mask | (
                    results["Product - Series"].str.contains(term, case=False, na=False) |
                    results["BUG headline"].str.contains(term, case=False, na=False)
                )
            results = results[mask]

    if version:
        mask = results.apply(
            lambda row: version_affects_bug(
                row["Known Affected Release(s)"], row.get("Known Fixed Releases"), version
            ),
            axis=1,
        )
        results = results[mask]

    if ios_version:
        mask = results.apply(
            lambda row: version_affects_bug(
                row["Known Affected Release(s)"], row.get("Known Fixed Releases"), ios_version
            ),
            axis=1,
        )
        results = results[mask]

    if sort_by == "severity":
        results = results.sort_values("Bug Severity", ascending=True)
    elif sort_by == "last_modified":
        results = results.sort_values("Last Modified", ascending=False)
    elif sort_by == "bug_id":
        results = results.sort_values("BUG Id")

    return results


def list_affected_releases(df):
    """CSV 内に登場する全 Known Affected Release(s) のユニーク一覧（降順ソート）"""
    all_releases = set()
    for releases in df["Known Affected Release(s)"].dropna():
        for release in str(releases).split():
            all_releases.add(release.strip())
    return sorted([r for r in all_releases if r], reverse=True)


# ==================== Excel レポート生成 ====================

def _write_simple_sheet(wb, sheet_name, headers, rows):
    """
    見出し行 + 単純な表形式データを1シートとして追加する共通ヘルパー。
    Palo Alto の CVE 検索結果や YAMAHA 等の貼り付け解析結果など、
    Cisco 固有ではないデータをシート分けしてExcelに含めるのに使う。

    Args:
        wb: openpyxl の Workbook
        sheet_name: シート名（31文字を超える場合は自動的に切り詰める）
        headers: 列見出しのリスト
        rows: 各行のリストのリスト（列数は headers と揃える）
    """
    # Excel のシート名は31文字までという制約があるため切り詰める
    ws = wb.create_sheet(sheet_name[:31])

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row_values in enumerate(rows, 2):
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 32

    return ws


def create_excel_report(results, analysis_data, search_params, include_release_notes=False,
                         translation_engine='google', deepl_api_key=None, nvidia_api_key=None,
                         extra_sheets=None, groq_api_key=None, gemini_api_key=None,
                         open_router_api_key=None, target_version=None):
    """
    Excel 形式のレポートを生成

    複数シート構成:
    - Sheet1 検索結果: バグ情報 + 分析結果（include_release_notes=True で症状/条件/回避策/詳細説明を日本語で追加）
    - Sheet2 分析詳細: 分析結果のみ
    - Sheet3 検索パラメータ: 検索条件とメタデータ
    - extra_sheets で指定した分、追加のシート（例: Palo Alto の CVE検索結果、
      YAMAHA等の貼り付け解析結果）

    Args:
        extra_sheets: [{"name": str, "headers": [...], "rows": [[...], ...]}, ...] の形式で
            追加シートを指定する（省略可）。複数ベンダーの結果を1つのExcelにまとめたい場合に使う。
        include_release_notes: True にすると、results の各行の "Release Note Enclosure" を
            parse_release_note() で解析し、翻訳した症状/条件/回避策/詳細説明を列として追加する。
            件数が多いと翻訳API呼び出しが増えて生成に時間がかかる点に注意。
        translation_engine, deepl_api_key, nvidia_api_key: include_release_notes=True の際に使う翻訳設定
        groq_api_key, gemini_api_key, open_router_api_key: 指定があれば、症状/回避策/詳細説明を
            翻訳の代わりに AI で日本語要約する（生ログ・スタックトレース等を除いた要点のみ）。
            いずれも無ければ通常の翻訳結果を使う。
        target_version: 指定すると "{target_version}影響" 列を追加し、そのバージョンへの
            影響有無を version_affects_bug() で判定して記載する（省略可）。
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "検索結果"
    ws2 = wb.create_sheet("分析詳細")
    ws3 = wb.create_sheet("検索パラメータ")

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet1: 検索結果
    # 深刻度順（Severity 数字の小さい順 = 重大な順）に並べ替えてから出力する。
    # 数字以外や欠損は末尾に回す
    sorted_results = results.copy()
    sorted_results["_severity_sort"] = pd.to_numeric(sorted_results.get("Bug Severity"), errors="coerce")
    sorted_results = sorted_results.sort_values(by="_severity_sort", ascending=True, na_position="last")

    headers = ["Bug ID", "Headline (日本語)", "Headline (英語原文・参考)", "Severity", "Status",
               "Affected Releases", "Fixed Releases", "利用機能", "素因"]
    if target_version:
        headers.append(f"{target_version}影響")
    headers += ["発生可能性", "発生しやすさ(推定)", "関連機能", "コメント", "URL"]
    release_note_cols = ["症状 (日本語)", "条件 (日本語)", "回避策 (日本語)", "詳細説明 (日本語)"]
    release_note_keys = ["症状", "条件", "回避策", "詳細説明"]
    if include_release_notes:
        headers = headers + release_note_cols + ["リリースノート原文（英語・参考）"]

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row, (idx, bug_row) in enumerate(sorted_results.iterrows(), 2):
        bug_id = bug_row["BUG Id"]
        analysis = analysis_data.get(bug_id, {})
        headline_en = bug_row.get("BUG headline", "")
        headline = bug_row.get("BUG headline (日本語)")
        if headline is None or (isinstance(headline, float) and pd.isna(headline)):
            headline = headline_en

        target_affected = None
        row_data = [
            bug_id,
            headline,
            headline_en,
            bug_row["Bug Severity"],
            bug_row["Bug Status"],
            bug_row["Known Affected Release(s)"],
            bug_row["Known Fixed Releases"],
            classify_bug_feature(headline_en),
            classify_bug_symptom(headline_en),
        ]
        if target_version:
            target_affected = version_affects_bug(
                bug_row["Known Affected Release(s)"], bug_row.get("Known Fixed Releases"), target_version
            )
            row_data.append("影響あり" if target_affected else "対象外")
        row_data += [
            analysis.get("possibility", "-"),
            estimate_occurrence_likelihood(bug_row["Bug Status"], headline_en, target_affected=target_affected),
            ", ".join(analysis.get("tags", [])),
            analysis.get("comment", ""),
            bug_row.get("URL", ""),
        ]

        if include_release_notes:
            raw_note = bug_row.get("Release Note Enclosure", "")
            if pd.isna(raw_note):
                raw_note = ""
            sections = None
            for key, col_name in zip(release_note_keys, release_note_cols):
                # 呼び出し側（app.py）で既に日本語訳/AI要約済みの列があれば、それをそのまま使い、
                # 二重に翻訳APIを呼ばない。無ければ（条件列、または results にまだ無い場合）
                # ここで初めて parse_release_note + 翻訳を行う
                existing = bug_row.get(col_name)
                if existing is not None and not (isinstance(existing, float) and pd.isna(existing)) and existing != "":
                    row_data.append(existing)
                    continue

                if sections is None:
                    sections = parse_release_note(raw_note)
                content = sections.get(key, "")
                if content:
                    summary = None
                    if groq_api_key or gemini_api_key or open_router_api_key:
                        summary = summarize_technical_text_ja(
                            content, groq_key=groq_api_key, gemini_key=gemini_api_key,
                            open_router_key=open_router_api_key
                        )
                    content = summary or translate_headline(
                        content, engine=translation_engine,
                        deepl_api_key=deepl_api_key, nvidia_api_key=nvidia_api_key
                    )
                row_data.append(content)
            row_data.append(clean_html_tags(raw_note))

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    _col_widths = {
        "Bug ID": 15, "Headline (日本語)": 40, "Headline (英語原文・参考)": 40, "Severity": 10,
        "Status": 12, "Affected Releases": 25, "Fixed Releases": 25, "利用機能": 22, "素因": 18,
        "発生可能性": 12, "発生しやすさ(推定)": 42, "関連機能": 20, "コメント": 30, "URL": 30,
        "症状 (日本語)": 35, "条件 (日本語)": 35, "回避策 (日本語)": 35, "詳細説明 (日本語)": 35,
        "リリースノート原文（英語・参考）": 45,
    }
    for col_idx, header in enumerate(headers, 1):
        width = 15 if header.endswith("影響") else _col_widths.get(header, 20)
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # Sheet2: 分析詳細
    analysis_headers = ["Bug ID", "発生可能性", "関連機能", "コメント"]

    for col, header in enumerate(analysis_headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row, (bug_id, data) in enumerate(analysis_data.items(), 2):
        row_data = [
            bug_id,
            data.get("possibility", "-"),
            ", ".join(data.get("tags", [])),
            data.get("comment", "")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 40

    # Sheet3: 検索パラメータ
    params_data = [
        ["検索パラメータ", ""],
        ["", ""],
        ["タイムスタンプ", datetime.now().isoformat()],
        ["機能", search_params.get("feature") or "（なし）"],
        ["バージョン", search_params.get("version") or "（なし）"],
        ["Severity フィルタ", ", ".join(map(str, search_params.get("severity", [])))],
        ["", ""],
        ["分析情報", ""],
        ["分析済みバグ数", len(analysis_data)],
        ["", ""],
        ["注意事項", ""],
        ["バージョン検索について",
         "旧バージョンから未修正のまま続いている可能性があるバグを推測で含みます。"
         "中間バージョンで実際に発生するとは限らないため、重要な判断の前に各バグページで確認してください。"],
    ]

    for row, (key, value) in enumerate(params_data, 1):
        cell_key = ws3.cell(row=row, column=1)
        cell_value = ws3.cell(row=row, column=2)

        cell_key.value = key
        cell_value.value = value
        cell_value.alignment = Alignment(wrap_text=True, vertical="top")

        if key in ["検索パラメータ", "分析情報", "注意事項"]:
            cell_key.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell_key.font = Font(bold=True)

        cell_key.border = border
        cell_value.border = border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 60

    for sheet_spec in (extra_sheets or []):
        _write_simple_sheet(wb, sheet_spec["name"], sheet_spec["headers"], sheet_spec["rows"])

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()


def create_combined_excel_report(cisco=None, extra_sheets=None):
    """
    Cisco（構造化検索結果）と、Palo Alto の CVE 検索結果・YAMAHA 等の貼り付け解析結果
    （どちらも extra_sheets 側）を、1つのExcelファイルにシート分けしてまとめる。
    どちらも省略可（両方 None なら空のExcelを返す）。

    Args:
        cisco: None、または以下の形式の dict:
            {"results": DataFrame, "analysis_data": dict, "search_params": dict,
             "include_release_notes": bool（省略可）, "translation_engine": str（省略可）,
             "deepl_api_key": str（省略可）, "nvidia_api_key": str（省略可）,
             "groq_api_key": str（省略可）, "gemini_api_key": str（省略可）,
             "open_router_api_key": str（省略可）, "target_version": str（省略可）}
        extra_sheets: [{"name": str, "headers": [...], "rows": [[...], ...]}, ...] または None

    Returns:
        Excel ファイルのバイト列
    """
    if cisco is not None:
        return create_excel_report(
            cisco["results"], cisco["analysis_data"], cisco["search_params"],
            include_release_notes=cisco.get("include_release_notes", False),
            translation_engine=cisco.get("translation_engine", "google"),
            deepl_api_key=cisco.get("deepl_api_key"), nvidia_api_key=cisco.get("nvidia_api_key"),
            groq_api_key=cisco.get("groq_api_key"), gemini_api_key=cisco.get("gemini_api_key"),
            open_router_api_key=cisco.get("open_router_api_key"),
            target_version=cisco.get("target_version"),
            extra_sheets=extra_sheets,
        )

    wb = Workbook()
    wb.remove(wb.active)  # デフォルトで作られる空シートを削除

    for sheet_spec in (extra_sheets or []):
        _write_simple_sheet(wb, sheet_spec["name"], sheet_spec["headers"], sheet_spec["rows"])

    if not wb.sheetnames:
        wb.create_sheet("結果なし")  # シートが1つも無いとExcelファイルとして不正になるため

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()


# ==================== NVD 検索（Cisco 以外のベンダー向け） ====================
#
# Cisco は bugSearch.csv という構造化データがあるため search_bugs() で高速に
# 検索できるが、Palo Alto や YAMAHA のようにそれに相当するデータセットを
# 持たないベンダーについては、NVD（米国立脆弱性データベース）の公開APIを
# キーワード検索することで代替する。CVE は全ベンダー共通でNVDに登録されており、
# CVSS スコア（＝緊急度）も構造化データとして取得できる。
# API仕様: https://nvd.nist.gov/developers/vulnerabilities

NVD_API_BASE = "https://services.nvd.nist.gov/rest/json/cves/2.0"

CVSS_SEVERITY_LABELS = {
    "CRITICAL": "緊急 (Critical)",
    "HIGH": "重要 (High)",
    "MEDIUM": "警告 (Medium)",
    "LOW": "注意 (Low)",
    "NONE": "情報 (None)",
}


def _extract_cvss(metrics):
    """CVE の metrics ブロックから CVSS スコアと深刻度を取り出す（v3.1 → v3.0 → v2 の優先順）"""
    for key in ("cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
        entries = metrics.get(key)
        if entries:
            cvss_data = entries[0].get("cvssData", {})
            score = cvss_data.get("baseScore")
            severity = entries[0].get("baseSeverity") or cvss_data.get("baseSeverity")
            return score, severity
    return None, None


AFFECTED_LABELS = {
    True: "影響あり",
    False: "対象外 / 修正済みの可能性",
    None: "判定不可（データ不足）",
}


def _version_in_cpe_match(target, cpe_match):
    """
    target: _parse_version_tuple() で得た (major, minor, patch) タプル
    cpe_match: NVD configurations 内の1つの cpeMatch エントリ

    target が cpe_match の示すバージョン範囲に含まれるか判定する。
    範囲指定（versionStart/EndIncluding/Excluding）が無い場合は、
    criteria 文字列自体に埋め込まれた具体的なバージョンと比較する。
    """
    start_inc = cpe_match.get("versionStartIncluding")
    start_exc = cpe_match.get("versionStartExcluding")
    end_inc = cpe_match.get("versionEndIncluding")
    end_exc = cpe_match.get("versionEndExcluding")

    if not any([start_inc, start_exc, end_inc, end_exc]):
        # 範囲指定なし → criteria (cpe:2.3:o:vendor:product:VERSION:...) 内の
        # バージョン部分（6番目のコロン区切り要素）と直接比較する
        parts = cpe_match.get("criteria", "").split(":")
        if len(parts) > 5:
            specific_version = parts[5]
            if specific_version not in ("*", "-", ""):
                v = _parse_version_tuple(specific_version)
                if v:
                    return v == target
        # 具体的なバージョンが取れない（ワイルドカードのみ）→ 判定材料なし
        return None

    if start_inc:
        v = _parse_version_tuple(start_inc)
        if v and target < v:
            return False
    if start_exc:
        v = _parse_version_tuple(start_exc)
        if v and target <= v:
            return False
    if end_inc:
        v = _parse_version_tuple(end_inc)
        if v and target > v:
            return False
    if end_exc:
        v = _parse_version_tuple(end_exc)
        if v and target >= v:
            return False

    return True


def check_version_affected(cve, target_version):
    """
    1件の NVD CVE 生データ（configurations を含む）に対して、
    target_version が影響を受けるかどうかを判定する。

    NVD の configurations には、CPE（対象製品・バージョン範囲）単位で
    「脆弱かどうか」と「どのバージョン範囲か」が構造化データとして入っている。
    Cisco の version_affects_bug() と同様の考え方で、指定バージョンが
    その範囲内かどうかをバージョン番号として比較する。

    Returns:
        True  : 影響を受ける可能性が高い（脆弱なバージョン範囲に含まれる）
        False : 影響を受けない（対象範囲外、＝修正済みの可能性）
        None  : configurations が無い、またはバージョン範囲が特定できず判定不可
    """
    target = _parse_version_tuple(target_version)
    if target is None:
        return None

    configurations = cve.get("configurations", [])
    if not configurations:
        return None

    saw_definitive_match = False
    for config in configurations:
        for node in config.get("nodes", []):
            for cpe_match in node.get("cpeMatch", []):
                if not cpe_match.get("vulnerable", True):
                    continue
                verdict = _version_in_cpe_match(target, cpe_match)
                if verdict is True:
                    return True
                if verdict is False:
                    saw_definitive_match = True

    if saw_definitive_match:
        return False
    return None


def search_cve_by_keyword(keyword, results_limit=20, api_key=None, timeout=20, target_version=None):
    """
    NVD の公開APIをキーワード検索し、該当する CVE の一覧を返す。

    Args:
        keyword: 検索キーワード（例: "PAN-OS 11.1.2", "Yamaha RTX830"）
        results_limit: 取得件数上限（NVD 側の resultsPerPage）
        api_key: NVD API キー（省略可。無しだと 5リクエスト/30秒 とレート制限が厳しい。
            https://nvd.nist.gov/developers/request-an-api-key から無料取得可能）
        timeout: リクエストタイムアウト秒数
        target_version: 指定すると、各 CVE が該当バージョンに影響するかどうかを
            NVD の configurations（CPEバージョン範囲）から判定して affected / affected_ja
            を結果に追加する（Cisco の version_affects_bug 相当）。数値として
            解釈できないバージョンを渡した場合や configurations が無い CVE では
            affected は None（判定不可）になる。

    Returns:
        成功時: [{"cve_id", "description_en", "cvss_score", "severity",
                  "severity_ja", "published", "url", ["affected", "affected_ja"]}, ...]
        失敗時: {"error": "エラーメッセージ"}
    """
    headers = {"apiKey": api_key} if api_key else {}
    params = {"keywordSearch": keyword, "resultsPerPage": results_limit}

    try:
        response = requests.get(NVD_API_BASE, params=params, headers=headers, timeout=timeout)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        return {"error": str(e)}

    return _parse_nvd_response(data, target_version=target_version)


def _parse_nvd_response(data, target_version=None):
    """NVD API v2.0 の生JSONレスポンスを扱いやすい辞書のリストに変換する"""
    results = []

    for item in data.get("vulnerabilities", []):
        cve = item.get("cve", {})
        cve_id = cve.get("id", "")

        descriptions = cve.get("descriptions", [])
        description_en = next((d["value"] for d in descriptions if d.get("lang") == "en"), "")

        score, severity = _extract_cvss(cve.get("metrics", {}))

        references = cve.get("references", [])
        url = references[0]["url"] if references else f"https://nvd.nist.gov/vuln/detail/{cve_id}"

        entry = {
            "cve_id": cve_id,
            "description_en": description_en,
            "cvss_score": score,
            "severity": severity,
            "severity_ja": CVSS_SEVERITY_LABELS.get(severity, severity or "不明"),
            "published": cve.get("published", ""),
            "url": url,
        }

        if target_version:
            affected = check_version_affected(cve, target_version)
            entry["affected"] = affected
            entry["affected_ja"] = AFFECTED_LABELS[affected]

        results.append(entry)

    return results


def search_cve_with_translation(keyword, engine='google', deepl_api_key=None, nvidia_api_key=None,
                                 results_limit=20, api_key=None, target_version=None):
    """
    search_cve_by_keyword() の結果に日本語訳（description_ja）を付けて返す。

    target_version を指定した場合は、影響あり（affected=True）のCVEを最優先、
    次に判定不可（None）、最後に対象外（False）の順に並べ、各グループ内では
    CVSS スコアの高い順にソートする。target_version を指定しない場合は
    単純に CVSS スコアの高い順（緊急度が高い順）にソートする。
    """
    results = search_cve_by_keyword(
        keyword, results_limit=results_limit, api_key=api_key, target_version=target_version
    )
    if isinstance(results, dict) and "error" in results:
        return results

    for r in results:
        r["description_ja"] = translate_headline(
            r["description_en"], engine=engine, deepl_api_key=deepl_api_key, nvidia_api_key=nvidia_api_key
        )

    if target_version:
        affected_rank = {True: 0, None: 1, False: 2}
        results.sort(key=lambda r: (affected_rank[r["affected"]], -(r["cvss_score"] or 0)))
    else:
        results.sort(key=lambda r: r["cvss_score"] or 0, reverse=True)

    return results


# ==================== ベンダー公式ドキュメントの手動貼り付け解析 ====================
#
# Palo Alto の "Known and Addressed Issues" ページのように、公式サイトが
# ボット対策で自動取得できないベンダーのバグ一覧に対応するための機能。
# ユーザーがブラウザで開いてコピーしたテキストを解析し、ID単位に分解・
# カテゴリ分け・翻訳する。NVD 検索（セキュリティCVEのみ）ではカバーできない、
# 一般的な不具合情報を扱う。


# 形式1: Palo Alto の Known Issues 形式。ID がそれだけで1行を占める
# （例: "PAN-332943" という行の次の行から説明文が始まる）
_ISSUE_ID_LINE_RE = re.compile(r'^([A-Z][A-Z0-9]*-\d+)\s*$', re.MULTILINE)

# 形式2: YAMAHA RT シリーズのリリースノート形式。角括弧の連番の直後（同じ行）に
# 説明文が続く（例: "[12] IKEv2で、鍵交換の始動パケットを受信しない機能を追加した。"）
_BRACKET_ITEM_RE = re.compile(r'^\[(\d+)\]\s*', re.MULTILINE)

# YAMAHA形式内のセクション見出し（例: "■バグ修正", "■新機能"）。
# 連番は見出しをまたいでリセットされる（新機能側の [1] とバグ修正側の [1] は別物）ため、
# どのセクション配下かを識別してIDの衝突を防ぐ
_SECTION_HEADER_RE = re.compile(r'^■\s*(.+?)\s*$', re.MULTILINE)

_WORKAROUND_RE = re.compile(r'Workaround\s*:\s*', re.IGNORECASE)
_WHITESPACE_RE = re.compile(r'\s+')


def _normalize_block_text(text):
    """
    複数行にまたがるブロックの空白を正規化する。

    改行を含む空白は、前後の文字が両方とも非ASCII（日本語等）の場合はスペースを
    挟まずに連結する（例: 行折り返しで分かれた「処理するよ」+「うにした」を
    「処理するように した」ではなく「処理するようにした」に戻す）。
    それ以外（英語の行折り返し等）は単一スペースに圧縮する。
    """
    text = text.strip()

    def _replace(m):
        ws = m.group(0)
        if '\n' not in ws:
            return ' '
        before = m.string[m.start() - 1] if m.start() > 0 else ''
        after = m.string[m.end()] if m.end() < len(m.string) else ''
        if before and after and not before.isascii() and not after.isascii():
            return ''
        return ' '

    return _WHITESPACE_RE.sub(_replace, text)


def parse_vendor_known_issues(raw_text):
    """
    ベンダー公式ページからコピーしたテキストを、課題単位のリストに分解する。
    2つの形式を自動判定する:

    1. Palo Alto 形式: "PAN-332943" のように ID が単独で1行を占め、
       次の行から説明文（+ 任意で "Workaround:" 以降）が続く
    2. YAMAHA 形式: "[12] 説明文..." のように角括弧の連番の直後に説明文が
       同じ行から続く（連番はそのリリースノート内でのみ一意）

    どちらのIDパターンにも一致しなければ空リストを返す。

    Args:
        raw_text: ブラウザからコピーした生テキスト

    Returns:
        [{"id": "PAN-332943", "description": "...", "workaround": "...", "section": None or "バグ修正"}, ...]
        YAMAHA形式の場合 id は "[12]" のようになり、workaround は常に空文字列
        （YAMAHAのリリースノートに Workaround の概念が無いため）。
        "■バグ修正" 等のセクション見出しが検出できた場合、その配下の項目には
        section にその見出し文字列が入る（無ければ None）。YAMAHA形式は見出しを
        またいで連番がリセットされる（新機能側の [1] とバグ修正側の [1] は別物）ため、
        同じ ID の項目を区別する際は id 単独ではなく section と組み合わせて使うこと。
    """
    if not raw_text or not raw_text.strip():
        return []

    matches = list(_ISSUE_ID_LINE_RE.finditer(raw_text))
    if matches:
        issues = []
        for i, m in enumerate(matches):
            issue_id = m.group(1)
            start = m.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(raw_text)
            block = raw_text[start:end].strip()

            wa_match = _WORKAROUND_RE.search(block)
            if wa_match:
                description = _normalize_block_text(block[:wa_match.start()])
                workaround = _normalize_block_text(block[wa_match.end():])
            else:
                description = _normalize_block_text(block)
                workaround = ""

            if description:
                issues.append({"id": issue_id, "description": description, "workaround": workaround, "section": None})
        return issues

    matches = list(_BRACKET_ITEM_RE.finditer(raw_text))
    if matches:
        section_headers = list(_SECTION_HEADER_RE.finditer(raw_text))

        def _section_at(pos):
            current = None
            for sm in section_headers:
                if sm.start() <= pos:
                    current = sm.group(1)
                else:
                    break
            return current

        issues = []
        for i, m in enumerate(matches):
            issue_id = f"[{m.group(1)}]"
            start = m.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(raw_text)
            description = _normalize_block_text(raw_text[start:end])

            if description:
                issues.append({
                    "id": issue_id,
                    "description": description,
                    "workaround": "",
                    "section": _section_at(m.start()),
                })
        return issues

    return []


# カテゴリ名 → 判定に使うキーワード（説明文の小文字化テキストに対して部分一致で判定）。
# 技術用語（ipsec, vpn 等）は日本語の文章中でも半角英数字で書かれることが多いため、
# 英語（Palo Alto）・日本語（YAMAHA）どちらの説明文にも同じキーワードセットで対応できる。
# 1件が複数カテゴリに該当することもある。どれにも該当しなければ「一般 / その他」
VENDOR_ISSUE_CATEGORY_KEYWORDS = {
    "Panorama": ["panorama", "m-700", "template", "log collector"],
    "HA / クラスタ": [
        "ha pair", "high availability", "cluster", "hsf", "hsci",
        " leader", "follower", "standalone", "split brain", "active/passive",
    ],
    "クラウド (GCP/Azure)": ["gcp", "azure", " nsi ", "gvnic", "vm-series", "hotplug"],
    "GlobalProtect": ["globalprotect", "portal", "gateway"],
    "5G / セルラー": [
        "cellular", "5g", "apn", "ztp", "fail-to-wire", "fail-open",
        "zero touch provisioning", "-r-poe", "410r", "450r",
    ],
    "VPN / IPsec": ["ipsec", "vpn", "ike", "l2tp", "pptp"],
    "スイッチ連携 (L2MS)": ["l2ms", "swx", "wlx", "スイッチ"],
}


def categorize_vendor_issue(description):
    """
    説明文（英語・日本語どちらも可）からキーワードマッチで該当カテゴリを判定する。
    複数該当する場合は全て返す。どれにも該当しなければ ["一般 / その他"]。
    """
    text_lower = f" {description.lower()} "
    categories = [
        category for category, keywords in VENDOR_ISSUE_CATEGORY_KEYWORDS.items()
        if any(kw in text_lower for kw in keywords)
    ]
    return categories or ["一般 / その他"]


# ==================== EOL（サポート終了）情報取得 ====================
#
# ベンダー公式のEOLページ（例: paloaltonetworks.com）は自動取得できないことが多いが、
# endoflife.date プロジェクトが GitHub 上で構造化データ（YAML）として公開しており、
# raw.githubusercontent.com は本セッションでもブロックされずアクセスできることを確認済み。

ENDOFLIFE_RAW_BASE = "https://raw.githubusercontent.com/endoflife-date/endoflife.date/master/products"


def _yaml_value_to_str(value):
    """YAML パース後の値（datetime.date や bool 等）を文字列 or None に正規化する"""
    if value is None or value is False:
        return None
    if isinstance(value, _date):
        return value.isoformat()
    return str(value)


def get_eol_info(product_slug, timeout=15):
    """
    endoflife.date（GitHub上のYAMLデータ）から、指定プロダクトのバージョン系統ごとの
    リリース日・EOL日・最新パッチ・関連リンクを取得する。

    Args:
        product_slug: endoflife.date 上のプロダクトスラッグ
            （例: "pan-os" = Palo Alto PAN-OS。一覧は https://endoflife.date/ 参照）
        timeout: リクエストタイムアウト秒数

    Returns:
        成功時: [{"release_cycle": "12.2", "release_date": "2026-07-30",
                  "eol": "2028-08-28" または None（現役でEOL未定）,
                  "is_eol": bool（今日時点でEOL済みか）,
                  "latest": "12.2.2", "latest_release_date": "...", "link": "https://..."}, ...]
            release_date の新しい順（降順）で返す
        失敗時: {"error": "エラーメッセージ"}
    """
    url = f"{ENDOFLIFE_RAW_BASE}/{product_slug}.md"

    try:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        raw = response.text
    except requests.exceptions.HTTPError as e:
        if e.response is not None and e.response.status_code == 404:
            return {"error": f"'{product_slug}' が見つかりません（例: Cisco IOS XE は 'cisco-ios-xe'）。"
                              "正しいスラッグは https://endoflife.date/ で確認してください。"}
        return {"error": str(e)}
    except Exception as e:
        return {"error": str(e)}

    parts = raw.split("---", 2)
    if len(parts) < 3:
        return {"error": f"'{product_slug}' の YAML frontmatter が見つかりませんでした"}

    try:
        data = yaml.safe_load(parts[1])
    except Exception as e:
        return {"error": f"YAML 解析エラー: {e}"}

    if not data or "releases" not in data:
        return {"error": f"'{product_slug}' の releases データが見つかりませんでした"}

    today = datetime.now().date()
    results = []

    for r in data["releases"]:
        eol_str = _yaml_value_to_str(r.get("eol"))
        is_eol = False
        if eol_str:
            try:
                is_eol = _date.fromisoformat(eol_str) <= today
            except ValueError:
                is_eol = False

        results.append({
            "release_cycle": str(r.get("releaseCycle", "")),
            "release_label": r.get("releaseLabel"),
            "release_date": _yaml_value_to_str(r.get("releaseDate")),
            "eol": eol_str,
            "is_eol": is_eol,
            "latest": r.get("latest"),
            "latest_release_date": _yaml_value_to_str(r.get("latestReleaseDate")),
            "link": r.get("link"),
        })

    results.sort(key=lambda r: r["release_date"] or "", reverse=True)
    return results


# ==================== Cisco 公式 EOL/EOS 通知の貼り付け解析 ====================
# endoflife.date の Cisco 製品データは実測ではなく「リリース日 + 固定オフセット」の
# 推定式であることがあり、Cisco 公式のEOL通知（例: IOS XE 17.17 の Last Date of
# Support が実際には2029-01-30なのに、endoflife.date 側は2026-03-31という推定式の
# 値になっていた）と大きくずれることが確認されている。Cisco公式のEOL/EOS通知ページ
# （"End-of-life milestones" の表）は自動取得できないため、貼り付けたテキストから
# マイルストーンと日付を抽出する。

_CISCO_EOL_DATE_RE = re.compile(
    r'(?:January|February|March|April|May|June|July|August|September|October|November|December)'
    r'\s+\d{1,2},\s*\d{4}'
)

# (内部キー, 表示名（日本語）, ラベルを検出する正規表現)
_CISCO_EOL_MILESTONE_PATTERNS = [
    ("announcement", "EOL発表日", r'End-of-Life Announcement Date'),
    ("end_of_sale", "販売終了日(EOS)", r'End-of-Sale Date'),
    ("last_ship", "最終出荷日", r'Last Ship Date'),
    ("end_of_sw_maintenance", "SWメンテナンスリリース終了日", r'End of SW Maintenance Releases Date'),
    ("end_of_security_support", "脆弱性/セキュリティサポート終了日", r'End of Vulnerability/Security Support'),
    ("end_of_service_contract_renewal", "サービス契約更新終了日", r'End of Service Contract Renewal'),
    ("last_date_of_support", "サポート終了日（最終・実質的なEOL）", r'Last Date of Support'),
]


def parse_cisco_eol_milestones(raw_text):
    """
    Cisco公式のEOL/EOS通知ページ（"End-of-life milestones" の表）からコピーした
    テキストを解析し、マイルストーン名と日付のペアを抽出する。

    "Last Date of Support"（サポート終了日）が実質的な最終EOL日にあたる。
    endoflife.date 等の推定値より、こちらの方が正確な公式情報である点に注意。

    Args:
        raw_text: Cisco公式ページからコピーした生テキスト

    Returns:
        [{"milestone": "サポート終了日（最終・実質的なEOL）", "date": "January 30, 2029"}, ...]
        （表示順は _CISCO_EOL_MILESTONE_PATTERNS の並び順）
        該当するマイルストーンが1つも見つからなければ空リストを返す
    """
    if not raw_text or not raw_text.strip():
        return []

    rows = []
    for key, label_ja, label_pattern in _CISCO_EOL_MILESTONE_PATTERNS:
        m = re.search(label_pattern, raw_text, re.IGNORECASE)
        if not m:
            continue
        tail = raw_text[m.end():m.end() + 1000]
        date_match = _CISCO_EOL_DATE_RE.search(tail)
        if date_match:
            rows.append({"milestone": label_ja, "date": date_match.group(0)})

    return rows


# NX-OS は IOS XE と違い、バージョン専用ページではなく「NX-OS EoL Milestones」という
# 全メジャーリリース分をまとめた1つの表（NX-OS Major Release / EoSWM Date / EoVSS/LDoS）
# で公開されている。日付表記も "Nov 30 2023" のようにカンマ無しの短縮月名。
_CISCO_NXOS_RELEASE_RE = re.compile(r'\d+\.\d+\(x\)')
_CISCO_NXOS_DATE_RE = re.compile(
    r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s*\d{4}',
    re.IGNORECASE
)


def parse_cisco_nxos_eol_table(raw_text):
    """
    Cisco NX-OS ソフトウェアの「NX-OS EoL Milestones」表（全メジャーリリースをまとめた
    一覧表）からコピーしたテキストを解析する。IOS XEの「1バージョン=1ページ」形式とは
    異なり、こちらは "10.2(x)" のようなメジャーリリース番号ごとに
    EoSWM（ソフトウェアメンテナンス終了日）と EoVSS/LDoS（脆弱性サポート終了日/
    最終サポート終了日、2つの日付が "/" 区切りで書かれることがある）が並ぶ。

    Returns:
        [{"release": "10.2(x)", "eoswm": "Nov 30 2023",
          "eovss_ldos": "Feb 28 2025/Aug 31 2025"}, ...]
        該当するリリース行が1つも見つからなければ空リストを返す
    """
    if not raw_text or not raw_text.strip():
        return []

    matches = list(_CISCO_NXOS_RELEASE_RE.finditer(raw_text))
    rows = []
    for i, m in enumerate(matches):
        release = m.group(0)
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else min(len(raw_text), start + 300)
        chunk = raw_text[start:end]
        dates = _CISCO_NXOS_DATE_RE.findall(chunk)
        if not dates:
            continue
        rows.append({
            "release": release,
            "eoswm": dates[0],
            "eovss_ldos": "/".join(dates[1:]),
        })

    return rows


# Cisco公式のEOL/EOS通知ページから貼り付けて確認済みの実データ。
# endoflife.date は Cisco 製品の EOL を推定式で計算しており不正確なことがあるため、
# 貼り付け解析で検証済みの値をここに蓄積し、OS名+バージョン入力だけで即座に参照できるようにする。
CISCO_EOL_KNOWN_DATA = {
    "cisco-ios-xe": {
        "17.14": {
            "milestones": [
                {"milestone": "EOL発表日", "date": "August 30, 2024"},
                {"milestone": "販売終了日(EOS)", "date": "August 30, 2024"},
                {"milestone": "最終出荷日", "date": "November 28, 2024"},
                {"milestone": "SWメンテナンスリリース終了日", "date": "August 30, 2025"},
                {"milestone": "脆弱性/セキュリティサポート終了日", "date": "August 30, 2025"},
                {"milestone": "サービス契約更新終了日", "date": "August 30, 2028"},
                {"milestone": "サポート終了日（最終・実質的なEOL）", "date": "August 30, 2028"},
            ],
            "note": None,
        },
        "17.15": {
            "milestones": [
                {"milestone": "EOL発表日", "date": "October 28, 2024"},
                {"milestone": "販売終了日(EOS)", "date": "October 28, 2024"},
                {"milestone": "最終出荷日", "date": "January 26, 2025"},
                {"milestone": "SWメンテナンスリリース終了日", "date": "October 28, 2025"},
                {"milestone": "脆弱性/セキュリティサポート終了日", "date": "October 28, 2025"},
                {"milestone": "サービス契約更新終了日", "date": "October 28, 2028"},
                {"milestone": "サポート終了日（最終・実質的なEOL）", "date": "October 28, 2028"},
            ],
            "note": "Cisco は 17.18 への移行を推奨（マイグレーション対象）。",
        },
        "17.16": {
            "milestones": [
                {"milestone": "EOL発表日", "date": "January 30, 2025"},
                {"milestone": "販売終了日(EOS)", "date": "January 30, 2025"},
                {"milestone": "最終出荷日", "date": "April 30, 2025"},
                {"milestone": "SWメンテナンスリリース終了日", "date": "January 30, 2026"},
                {"milestone": "脆弱性/セキュリティサポート終了日", "date": "January 30, 2026"},
                {"milestone": "サービス契約更新終了日", "date": "January 30, 2029"},
                {"milestone": "サポート終了日（最終・実質的なEOL）", "date": "January 30, 2029"},
            ],
            "note": None,
        },
        "17.17": {
            "milestones": [
                {"milestone": "EOL発表日", "date": "March 31, 2025"},
                {"milestone": "販売終了日(EOS)", "date": "March 31, 2025"},
                {"milestone": "最終出荷日", "date": "June 29, 2025"},
                {"milestone": "SWメンテナンスリリース終了日", "date": "March 31, 2026"},
                {"milestone": "脆弱性/セキュリティサポート終了日", "date": "March 31, 2026"},
                {"milestone": "サービス契約更新終了日", "date": "January 30, 2029"},
                {"milestone": "サポート終了日（最終・実質的なEOL）", "date": "January 30, 2029"},
            ],
            "note": None,
        },
    },
    "cisco-nx-os": {
        "10.2": {"eoswm": "Nov 30 2023", "eovss_ldos": "Nov 30 2025"},
        "10.3": {"eoswm": "May 31 2024", "eovss_ldos": "May 31 2026"},
        "10.4": {"eoswm": "Nov 30 2024", "eovss_ldos": "Nov 30 2026"},
        "10.5": {"eoswm": "May 31 2025", "eovss_ldos": "May 31 2027"},
        "10.6": {"eoswm": "Nov 30 2025", "eovss_ldos": "Nov 30 2027"},
        "10.7": {"eoswm": "May 31 2026", "eovss_ldos": "May 31 2028"},
    },
}


def _normalize_cisco_eol_version(version):
    """バージョン表記のゆらぎ（末尾の "(x)"、前後空白、大文字小文字）を吸収する。"""
    if not version:
        return ""
    v = version.strip().lower()
    v = re.sub(r'\(x\)\s*$', '', v)
    return v.strip()


def lookup_cisco_eol(os_family, version):
    """OS名(cisco-ios-xe / cisco-nx-os)とバージョンから、貼り付け解析で検証済みの
    EOLデータを直接返す。データが無ければ None を返す（呼び出し側で貼り付け解析への
    誘導メッセージを表示する）。"""
    if not os_family or not version:
        return None
    family_data = CISCO_EOL_KNOWN_DATA.get(os_family)
    if not family_data:
        return None
    key = _normalize_cisco_eol_version(version)
    for known_version, data in family_data.items():
        if _normalize_cisco_eol_version(known_version) == key:
            return {"os_family": os_family, "version": known_version, **data}
    return None
