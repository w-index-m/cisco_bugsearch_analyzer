import streamlit as st
import pandas as pd
import io
import re
import json
from html import unescape
from datetime import datetime
from deep_translator import GoogleTranslator

try:
    import deepl
    DEEPL_AVAILABLE = True
except ImportError:
    DEEPL_AVAILABLE = False

import requests
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

st.set_page_config(page_title="Cisco Bug Search Analyzer", layout="wide")

st.title("🔍 Cisco Bug Search Analyzer")
st.markdown("Cisco バグ検索システム - 機能とバージョンから該当するバグを検索")

def clean_html_tags(text):
    """HTML タグを削除して日本語対応テキストに変換"""
    if not text:
        return ""
    text = re.sub(r'<[^>]+>', '', text)
    text = unescape(text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def parse_release_note(note_text):
    """リリースノートをセクション別に解析"""
    if not note_text:
        return {}

    text = clean_html_tags(note_text)
    sections = {}

    section_mapping = {
        'Symptom': '症状',
        'Symptôme': '症状',
        'Conditions': '条件',
        'Conditions d\'activation': '条件',
        'Workaround': '回避策',
        'Contournement': '回避策',
        'Further Problem Description': '詳細説明',
        'Description additionnelle du problème': '詳細説明'
    }

    for eng_key, jp_key in section_mapping.items():
        pattern = f'{eng_key}[:\s]*'
        if re.search(pattern, text, re.IGNORECASE):
            parts = re.split(pattern, text, flags=re.IGNORECASE, maxsplit=1)
            if len(parts) > 1:
                content = parts[1].split('\n')[0][:200]
                sections[jp_key] = content

    return sections

@st.cache_data
def translate_headline_deepl(text, api_key):
    """DeepL API を使用して日本語に翻訳"""
    if not text or len(text) < 3:
        return text
    try:
        translator = deepl.Translator(api_key)
        result = translator.translate_text(text, source_lang="EN", target_lang="JA")
        return result.text
    except Exception as e:
        return None

@st.cache_data
def translate_headline_google(text):
    """Google Translate を使用して日本語に翻訳"""
    if not text or len(text) < 3:
        return text
    try:
        translator = GoogleTranslator(source_language='en', target_language='ja')
        return translator.translate(text)
    except Exception as e:
        return None

def translate_headline_nvidia(text, api_key):
    """NVIDIA Riva Translation を使用して日本語に翻訳"""
    if not text or len(text) < 3:
        return None
    if not api_key:
        return None

    try:
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }

        payload = {
            "model": "nvidia/riva-translate-4b-instruct-v2",
            "messages": [
                {
                    "role": "user",
                    "content": f"Translate the following text from English to Japanese. Only return the translated text without any explanation.\n\nText: {text}"
                }
            ],
            "temperature": 0.5,
            "top_p": 0.9,
            "max_tokens": 512
        }

        response = requests.post(
            "https://integrate.api.nvidia.com/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=10
        )

        if response.status_code == 200:
            result = response.json()
            if "choices" in result and len(result["choices"]) > 0:
                return result["choices"][0]["message"]["content"].strip()
        return None

    except Exception as e:
        return None

def translate_headline(text, engine='google', deepl_api_key=None, nvidia_api_key=None):
    """翻訳エンジンを指定してヘッドラインを翻訳"""
    if not text or len(text) < 3:
        return text

    if engine == 'nvidia' and nvidia_api_key:
        result = translate_headline_nvidia(text, nvidia_api_key)
        if result:
            return result

    if engine == 'deepl' and deepl_api_key and DEEPL_AVAILABLE:
        result = translate_headline_deepl(text, deepl_api_key)
        if result:
            return result

    result = translate_headline_google(text)
    if result:
        return result

    return text

def save_analysis_to_json(analysis_data):
    """分析結果を JSON に変換"""
    output = {
        "timestamp": datetime.now().isoformat(),
        "analysis_count": len(analysis_data),
        "bugs": []
    }

    for bug_id, data in analysis_data.items():
        output["bugs"].append({
            "bug_id": bug_id,
            "possibility": data.get("possibility", "-"),
            "tags": data.get("tags", []),
            "comment": data.get("comment", "")
        })

    return json.dumps(output, ensure_ascii=False, indent=2)

def load_analysis_from_json(json_str):
    """JSON から分析結果を読み込み"""
    try:
        data = json.loads(json_str)
        analysis = {}
        for bug in data.get("bugs", []):
            analysis[bug["bug_id"]] = {
                "possibility": bug.get("possibility", "Medium"),
                "tags": bug.get("tags", []),
                "comment": bug.get("comment", "")
            }
        return analysis
    except Exception as e:
        st.error(f"JSON 読み込みエラー: {e}")
        return {}

def assess_bug_with_groq(headline, release_note, user_comment, api_key):
    """Groq API を使用して発生可能性を判定"""
    if not api_key:
        return None

    try:
        client = Groq(api_key=api_key)
        prompt = f"""以下のバグ情報から、発生の可能性を High/Medium/Low で判定してください。

【バグタイトル（日本語）】
{headline}

【リリースノート】
{release_note[:500]}

【ユーザーコメント】
{user_comment if user_comment else "（コメントなし）"}

【判定フォーマット】
可能性: [High/Medium/Low]
理由: [簡潔な理由]

日本語で回答してください。"""

        message = client.messages.create(
            model="mixtral-8x7b-32768",
            max_tokens=200,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        return message.content[0].text
    except Exception as e:
        return None

def assess_bug_with_gemini(headline, release_note, user_comment, api_key):
    """Gemini API を使用して発生可能性を判定"""
    if not api_key or not GEMINI_AVAILABLE:
        return None

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-pro')

        prompt = f"""以下のバグ情報から、発生の可能性を High/Medium/Low で判定してください。

【バグタイトル（日本語）】
{headline}

【リリースノート】
{release_note[:500]}

【ユーザーコメント】
{user_comment if user_comment else "（コメントなし）"}

【判定フォーマット】
可能性: [High/Medium/Low]
理由: [簡潔な理由]

日本語で回答してください。"""

        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return None

def assess_bug_with_open_router(headline, release_note, user_comment, api_key):
    """Open Router API を使用して発生可能性を判定"""
    if not api_key:
        return None

    try:
        prompt = f"""以下のバグ情報から、発生の可能性を High/Medium/Low で判定してください。

【バグタイトル（日本語）】
{headline}

【リリースノート】
{release_note[:500]}

【ユーザーコメント】
{user_comment if user_comment else "（コメントなし）"}

【判定フォーマット】
可能性: [High/Medium/Low]
理由: [簡潔な理由]

日本語で回答してください。"""

        response = requests.post(
            url="https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "HTTP-Referer": "https://streamlit.io",
                "X-Title": "Cisco Bug Search Analyzer"
            },
            json={
                "model": "mistralai/mistral-7b-instruct",
                "messages": [
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 200
            },
            timeout=10
        )

        if response.status_code == 200:
            result = response.json()
            if "choices" in result and len(result["choices"]) > 0:
                return result["choices"][0]["message"]["content"]
        return None
    except Exception as e:
        return None

def assess_bug_possibility(headline_ja, release_note, user_comment, groq_key=None, gemini_key=None, open_router_key=None):
    """
    複数の AI を試してバグの発生可能性を判定

    フォールバック順:
    1. Groq
    2. Gemini
    3. Open Router
    """
    assessments = [
        ("Groq", assess_bug_with_groq(headline_ja, release_note, user_comment, groq_key)),
        ("Gemini", assess_bug_with_gemini(headline_ja, release_note, user_comment, gemini_key)),
        ("Open Router", assess_bug_with_open_router(headline_ja, release_note, user_comment, open_router_key))
    ]

    for engine_name, result in assessments:
        if result:
            return {"engine": engine_name, "result": result}

    return None

def create_excel_report(results, analysis_data, search_params):
    """
    Excel 形式のレポートを生成

    複数シート構成:
    - Sheet1: 検索結果
    - Sheet2: 分析詳細
    - Sheet3: 検索パラメータ
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "検索結果"
    ws2 = wb.create_sheet("分析詳細")
    ws3 = wb.create_sheet("検索パラメータ")

    # スタイル定義
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet1: 検索結果
    headers = ["Bug ID", "Headline (日本語)", "Severity", "Status",
               "Affected Releases", "Fixed Releases", "発生可能性", "関連機能", "コメント"]

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row, (idx, bug_row) in enumerate(results.iterrows(), 2):
        bug_id = bug_row["BUG Id"]
        analysis = analysis_data.get(bug_id, {})

        row_data = [
            bug_id,
            bug_row["BUG headline (日本語)"],
            bug_row["Bug Severity"],
            bug_row["Bug Status"],
            bug_row["Known Affected Release(s)"],
            bug_row["Known Fixed Releases"],
            analysis.get("possibility", "-"),
            ", ".join(analysis.get("tags", [])),
            analysis.get("comment", "")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # 列幅設定
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 25
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 20
    ws1.column_dimensions['I'].width = 30

    # Sheet2: 分析詳細
    analysis_headers = ["Bug ID", "発生可能性", "関連機能", "コメント"]

    for col, header in enumerate(analysis_headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row, (bug_id, data) in enumerate(analysis_data.items(), 2):
        row_data = [
            bug_id,
            data.get("possibility", "-"),
            ", ".join(data.get("tags", [])),
            data.get("comment", "")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 40

    # Sheet3: 検索パラメータ
    params_data = [
        ["検索パラメータ", ""],
        ["", ""],
        ["タイムスタンプ", datetime.now().isoformat()],
        ["機能", search_params.get("feature", "（なし）")],
        ["バージョン", search_params.get("version", "（なし）")],
        ["Severity フィルタ", ", ".join(map(str, search_params.get("severity", [])))],
        ["", ""],
        ["分析情報", ""],
        ["分析済みバグ数", len(analysis_data)],
    ]

    for row, (key, value) in enumerate(params_data, 1):
        cell_key = ws3.cell(row=row, column=1)
        cell_value = ws3.cell(row=row, column=2)

        cell_key.value = key
        cell_value.value = value

        if key in ["検索パラメータ", "分析情報"]:
            cell_key.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell_key.font = Font(bold=True)

        cell_key.border = border
        cell_value.border = border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 40

    # Excel ファイルをバイト列に変換
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()

def get_cisco_release_notes_url(product, version):
    """Cisco 公式リリースノート URL を生成"""
    product_lower = product.lower()
    if 'catalyst 9200' in product_lower or '9200' in product_lower:
        version_short = version.replace('.', '-')
        return f"https://www.cisco.com/c/en/us/td/docs/switches/lan/catalyst9200/software/release/{version_short}/release_notes/ol-{version_short}-9200.html"
    elif 'catalyst 9300' in product_lower or '9300' in product_lower:
        version_short = version.replace('.', '-')
        return f"https://www.cisco.com/c/en/us/td/docs/switches/lan/catalyst9300/software/release/{version_short}/release_notes/ol-{version_short}-9300.html"
    elif 'catalyst 9400' in product_lower or '9400' in product_lower:
        version_short = version.replace('.', '-')
        return f"https://www.cisco.com/c/en/us/td/docs/switches/lan/catalyst9400/software/release/{version_short}/release_notes/ol-{version_short}-9400.html"
    else:
        return None

@st.cache_data
def load_csv(file_path):
    """CSV ファイルを読み込み"""
    try:
        df = pd.read_csv(file_path)
        df.columns = df.columns.str.strip()
        return df
    except Exception as e:
        st.error(f"ファイル読み込みエラー: {e}")
        return None

# CSV ファイル読み込み（デフォルト）
default_csv = "bugSearch.csv"

# ファイルアップロード or デフォルトを使用
uploaded_file = st.file_uploader("CSV ファイルをアップロード（オプション）", type=["csv"])

if uploaded_file is not None:
    df = pd.read_csv(uploaded_file)
    df.columns = df.columns.str.strip()
else:
    df = load_csv(default_csv)

if df is not None:
    st.success(f"✓ {len(df)} 件のバグ情報を読み込みました")

    if "bug_analysis" not in st.session_state:
        st.session_state.bug_analysis = {}

    st.markdown("---")

    with st.expander("💾 分析データの管理"):
        col1, col2 = st.columns(2)

        with col1:
            st.write("**分析データをアップロード**")
            uploaded_json = st.file_uploader(
                "JSON ファイルを選択",
                type=["json"],
                key="analysis_uploader"
            )
            if uploaded_json is not None:
                json_content = uploaded_json.read().decode('utf-8')
                loaded_analysis = load_analysis_from_json(json_content)
                if loaded_analysis:
                    st.session_state.bug_analysis.update(loaded_analysis)
                    st.success(f"✓ {len(loaded_analysis)} 件の分析データを読み込みました")

        with col2:
            st.write("**分析データをダウンロード**")
            analysis_json = save_analysis_to_json(st.session_state.bug_analysis)
            st.download_button(
                label="📥 分析結果を JSON でダウンロード",
                data=analysis_json,
                file_name=f"bug_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )

    st.markdown("---")

    col1, col2 = st.columns([2, 1])
    with col1:
        st.subheader("IOS バージョンから検索")
    with col2:
        st.subheader("翻訳設定 & Severity")

    col1, col2 = st.columns([2, 1])
    with col2:
        st.markdown("**翻訳エンジン**")
        translation_engine = st.radio(
            "翻訳エンジン",
            ["Google", "DeepL", "NVIDIA Riva"],
            horizontal=False,
            index=0,
            label_visibility="collapsed"
        )

        deepl_api_key = None
        nvidia_api_key = None

        if translation_engine == "DeepL":
            if DEEPL_AVAILABLE:
                deepl_api_key = st.text_input(
                    "DeepL API キー",
                    type="password",
                    placeholder="API キーを入力"
                )
                if not deepl_api_key:
                    st.warning("DeepL API キーを入力してください")
            else:
                st.warning("deepl ライブラリがインストールされていません")
                translation_engine = "Google"

        elif translation_engine == "NVIDIA Riva":
            nvidia_api_key = st.text_input(
                "NVIDIA API キー",
                type="password",
                placeholder="API キーを入力"
            )
            if not nvidia_api_key:
                st.warning("NVIDIA API キーを入力してください")

        st.markdown("**Severity フィルタ**")
        severity_filter = st.multiselect(
            "Severity を選択",
            options=[1, 2, 3, 4, 5],
            default=[1, 2, 3],
            label_visibility="collapsed"
        )

        st.markdown("**AI 分析エンジン**")
        use_ai_analysis = st.checkbox("AI による可能性判定を使用", value=False)

        groq_api_key = None
        gemini_api_key = None
        open_router_api_key = None

        if use_ai_analysis:
            st.markdown("API キーを入力（持っているもののみ）:")
            groq_api_key = st.text_input(
                "Groq API キー",
                type="password",
                placeholder="gsk_...",
                label_visibility="collapsed"
            )
            gemini_api_key = st.text_input(
                "Gemini API キー",
                type="password",
                placeholder="AIza...",
                label_visibility="collapsed"
            )
            open_router_api_key = st.text_input(
                "Open Router キー",
                type="password",
                placeholder="sk-or-...",
                label_visibility="collapsed"
            )

    st.markdown("---")
    st.subheader("IOS バージョンから検索")

    all_affected_releases = set()
    for releases in df["Known Affected Release(s)"].dropna():
        for release in str(releases).split():
            all_affected_releases.add(release.strip())

    sorted_releases = sorted([r for r in all_affected_releases if r], reverse=True)

    selected_ios_version = st.selectbox(
        "IOS バージョンを選択",
        [""] + sorted_releases,
        format_func=lambda x: "バージョンを選択..." if x == "" else x
    )

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        feature = st.text_input(
            "機能を入力（Product / Headline）",
            placeholder="例：Catalyst 9300, multicast, etc."
        )

    with col2:
        version = st.text_input(
            "バージョンを入力（Known Affected Release）",
            placeholder="例：17.12.4, 17.15.2, etc."
        )

    search_button = st.button("🔎 バグを検索", type="primary")

    st.markdown("---")

    if selected_ios_version:
        st.info(f"📋 **IOS {selected_ios_version} のリリースノート情報**")

        ios_bugs = df[
            df["Known Affected Release(s)"].str.contains(selected_ios_version, case=False, na=False)
        ].copy()

        if len(ios_bugs) > 0:
            st.write(f"**このバージョンに影響するバグ: {len(ios_bugs)} 件**")

            with st.expander("🔍 このバージョンのリリースノート内のバグ情報"):
                for idx, bug in ios_bugs.head(10).iterrows():
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        st.write(f"**{bug['BUG Id']}** - {bug['BUG headline'][:60]}")
                    with col2:
                        severity = bug['Bug Severity']
                        st.write(f"Severity: {severity}")
                    with col3:
                        st.write(f"Status: {bug['Bug Status']}")

                if len(ios_bugs) > 10:
                    st.caption(f"... 他 {len(ios_bugs) - 10} 件")
        else:
            st.write("このバージョンのバグ情報がありません")

        st.markdown("---")

    if search_button or (feature or version):
        results = df.copy()

        if severity_filter:
            results = results[results["Bug Severity"].astype(int).isin(severity_filter)]

        if feature:
            mask = (
                results["Product - Series"].str.contains(feature, case=False, na=False) |
                results["BUG headline"].str.contains(feature, case=False, na=False)
            )
            results = results[mask]

        if version:
            mask = results["Known Affected Release(s)"].str.contains(version, case=False, na=False)
            results = results[mask]

        if selected_ios_version:
            mask = results["Known Affected Release(s)"].str.contains(selected_ios_version, case=False, na=False)
            results = results[mask]

        st.subheader(f"検索結果: {len(results)} 件")

        if len(results) > 0:
            sort_by = st.selectbox(
                "ソート順",
                ["Bug Severity (高い順)", "Last Modified (新しい順)", "Bug ID"]
            )

            if sort_by == "Bug Severity (高い順)":
                results = results.sort_values("Bug Severity", ascending=True)
            elif sort_by == "Last Modified (新しい順)":
                results = results.sort_values("Last Modified", ascending=False)
            else:
                results = results.sort_values("BUG Id")

            display_results = results.copy()
            display_results["BUG headline (日本語)"] = display_results["BUG headline"].apply(
                lambda x: translate_headline(x, engine=translation_engine, deepl_api_key=deepl_api_key, nvidia_api_key=nvidia_api_key)
            )

            display_cols = ["BUG Id", "BUG headline (日本語)", "Bug Severity", "Bug Status",
                          "Known Affected Release(s)", "Known Fixed Releases"]

            st.dataframe(
                display_results[display_cols],
                use_container_width=True,
                hide_index=True
            )

            st.markdown("### 詳細情報")
            selected_idx = st.selectbox(
                "詳細を見るバグを選択",
                range(len(results)),
                format_func=lambda x: f"{results.iloc[x]['BUG Id']} - {translate_headline(results.iloc[x]['BUG headline'], engine=translation_engine, deepl_api_key=deepl_api_key, nvidia_api_key=nvidia_api_key)}"
            )

            if selected_idx is not None:
                bug = results.iloc[selected_idx]

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Bug ID", bug["BUG Id"])
                with col2:
                    severity = int(bug["Bug Severity"]) if pd.notna(bug["Bug Severity"]) else 0
                    st.metric("重要度", severity)
                with col3:
                    st.metric("ステータス", bug["Bug Status"])

                headline_en = bug["BUG headline"]
                headline_ja = translate_headline(headline_en, engine=translation_engine, deepl_api_key=deepl_api_key, nvidia_api_key=nvidia_api_key)

                st.write("**タイトル（日本語）:**", headline_ja)
                st.caption(f"英語: {headline_en}")
                if translation_engine != "Google":
                    st.caption(f"翻訳エンジン: {translation_engine}")

                st.markdown("---")
                st.markdown("### 📊 バグ分析")

                bug_id = bug["BUG Id"]
                if bug_id not in st.session_state.bug_analysis:
                    st.session_state.bug_analysis[bug_id] = {
                        "possibility": "Medium",
                        "tags": [],
                        "comment": "",
                        "ai_assessment": None
                    }

                analysis = st.session_state.bug_analysis[bug_id]

                col1, col2 = st.columns(2)
                with col1:
                    if use_ai_analysis and (groq_api_key or gemini_api_key or open_router_api_key):
                        if st.button("🤖 AI で分析", key=f"ai_btn_{bug_id}"):
                            with st.spinner("AI が分析中..."):
                                assessment = assess_bug_possibility(
                                    headline_ja,
                                    release_note,
                                    analysis["comment"],
                                    groq_key=groq_api_key,
                                    gemini_key=gemini_api_key,
                                    open_router_key=open_router_api_key
                                )

                                if assessment:
                                    st.session_state.bug_analysis[bug_id]["ai_assessment"] = assessment
                                    st.success(f"✓ {assessment['engine']} で分析完了")
                                else:
                                    st.error("AI による分析に失敗しました。API キーを確認してください。")

                    if analysis["ai_assessment"]:
                        st.info(f"**AI 分析結果（{analysis['ai_assessment']['engine']}）:**\n\n{analysis['ai_assessment']['result']}")

                with col2:
                    pass

                st.markdown("---")
                col1, col2 = st.columns(2)
                with col1:
                    possibility = st.selectbox(
                        "発生の可能性",
                        ["Low", "Medium", "High"],
                        index=["Low", "Medium", "High"].index(analysis["possibility"]),
                        key=f"possibility_{bug_id}"
                    )
                    analysis["possibility"] = possibility

                with col2:
                    tags = st.multiselect(
                        "関連機能タグ",
                        ["VPN", "Routing", "Multicast", "Security", "DHCP", "Access List", "QoS", "OSPF", "BGP", "その他"],
                        default=analysis["tags"],
                        key=f"tags_{bug_id}"
                    )
                    analysis["tags"] = tags

                comment = st.text_area(
                    "このバグについてのコメント",
                    value=analysis["comment"],
                    height=80,
                    key=f"comment_{bug_id}"
                )
                analysis["comment"] = comment

                release_note = bug["Release Note Enclosure"]
                sections = parse_release_note(release_note)

                if sections:
                    st.markdown("### 📋 リリースノート情報")

                    for section_name, content in sections.items():
                        st.write(f"**{section_name}:**")
                        st.caption(content)

                    # Cisco 公式ドキュメントへのリンク
                    affected_releases = str(bug["Known Affected Release(s)"]).split()
                    if affected_releases:
                        first_version = affected_releases[0]
                        product = bug["Product - Series"].split(',')[0].strip()
                        release_url = get_cisco_release_notes_url(product, first_version)

                        if release_url:
                            st.info(
                                f"📚 **参考資料**: "
                                f"[Cisco 公式リリースノート - {product} {first_version}]({release_url})"
                            )

                st.markdown("---")
                st.write("**製品:**", bug["Product - Series"])
                st.write("**影響を受けるバージョン:**", bug["Known Affected Release(s)"])
                st.write("**修正バージョン:**", bug["Known Fixed Releases"])
                st.write("**最終更新:**", bug["Last Modified"])

                col1, col2 = st.columns(2)
                with col1:
                    st.write("**参照リンク:**")
                    st.write(f"[Cisco Bug Search で詳細を確認]({bug['URL']})")

                with st.expander("📄 リリースノート全体"):
                    clean_note = clean_html_tags(release_note)
                    st.text(clean_note)

            st.markdown("---")
            st.markdown("### 📥 結果のエクスポート")

            export_results = results[display_cols].copy()
            export_results["発生可能性"] = export_results["BUG Id"].apply(
                lambda x: st.session_state.bug_analysis.get(x, {}).get("possibility", "-")
            )
            export_results["関連機能"] = export_results["BUG Id"].apply(
                lambda x: ", ".join(st.session_state.bug_analysis.get(x, {}).get("tags", []))
            )
            export_results["コメント"] = export_results["BUG Id"].apply(
                lambda x: st.session_state.bug_analysis.get(x, {}).get("comment", "")
            )

            csv_buffer = io.StringIO()
            export_results.to_csv(csv_buffer, index=False)
            csv_data = csv_buffer.getvalue()

            # Excel エクスポート用パラメータ
            search_params = {
                "feature": feature,
                "version": version,
                "severity": severity_filter
            }

            # Excel ファイルを生成
            excel_data = create_excel_report(results, st.session_state.bug_analysis, search_params)

            col1, col2, col3, col4 = st.columns(4)

            with col1:
                st.download_button(
                    label="📊 Excel でダウンロード",
                    data=excel_data,
                    file_name=f"cisco_bug_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            with col2:
                st.download_button(
                    label="📄 CSV でダウンロード",
                    data=csv_data,
                    file_name=f"bug_search_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )

            with col3:
                analysis_json = save_analysis_to_json(st.session_state.bug_analysis)
                st.download_button(
                    label="📋 分析結果を JSON で",
                    data=analysis_json,
                    file_name=f"bug_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json"
                )

            with col4:
                combined_export = {
                    "timestamp": datetime.now().isoformat(),
                    "search_params": search_params,
                    "results": export_results.to_dict('records')
                }
                st.download_button(
                    label="📦 完全レポート",
                    data=json.dumps(combined_export, ensure_ascii=False, indent=2),
                    file_name=f"bug_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json"
                )
        else:
            st.warning("該当するバグが見つかりませんでした")
else:
    st.error("CSV ファイルを読み込めません")
